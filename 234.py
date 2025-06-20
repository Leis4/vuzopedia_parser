import os
import re
import pickle
import pandas as pd
import matplotlib.pyplot as plt

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from tqdm import tqdm
from colorama import Fore, init

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

init(autoreset=True)

# --- Браузер и куки ---
def save_cookies(driver, path="cookies.pkl"):
    with open(path, "wb") as f:
        pickle.dump(driver.get_cookies(), f)

def load_cookies(driver, path="cookies.pkl"):
    with open(path, "rb") as f:
        cookies = pickle.load(f)
        for cookie in cookies:
            driver.add_cookie(cookie)

def get_page_with_wait(url, timeout=10):
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    driver = webdriver.Chrome(options=options)
    driver.get("https://vuzopedia.ru/")

    if os.path.exists("cookies.pkl"):
        try:
            load_cookies(driver)
            driver.get(url)
        except Exception:
            print(Fore.RED + "⚠️ Не удалось загрузить cookies.")
            driver.quit()
            exit(1)
    else:
        print(Fore.YELLOW + "🔐 Войдите вручную в браузере и нажмите Enter.")
        input("После входа нажмите Enter…")
        save_cookies(driver)
        driver.get(url)

    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.zpForMabile"))
        )
    except Exception:
        print(Fore.RED + "⚠️ Элемент zpForMabile не найден вовремя.")

    html = driver.page_source
    driver.quit()
    return html

# ---- BALL
def parse_scores_from_block(soup):
    scores = {"budget": {}, "paid": {}}
    block = soup.find("div", class_="zpForMabile")
    if not block:
        return scores
    current = None
    for p in block.find_all("p"):
        txt = p.get_text(strip=True).lower()
        if "бюджет" in txt:
            current = "budget"
        elif "плат" in txt:
            current = "paid"
        else:
            m = re.match(r"(\d{4})[:\s]+(\d+)", txt)
            if m and current:
                y, v = m.groups()
                scores[current][y] = int(v)
    return scores

#------------------------------
def parse_program(url):
    html = get_page_with_wait(url)
    soup = BeautifulSoup(html, 'html.parser')
    txt = soup.get_text("\n")

    cols = [
        "Ссылка", "Вуз", "Город", "Название программы", "Код направления",
        "Количество бюджетных мест", "Количество платных мест",
        "Стоимость обучения", "Минимальный балл ЕГЭ",
        "Общежитие", "Военный учебный центр"
    ] + [f"Проходной балл {y}" for y in ["2020", "2021", "2022", "2023", "2024"]] + \
        ["BVI_count", "BVI_pct", "Target_count", "Target_score",
         "SVO_count", "SVO_score", "Special_count", "Special_score"]
    data = dict.fromkeys(cols, "")

    data["Ссылка"] = url
    h1 = soup.find("h1")
    if h1: data["Название программы"] = h1.text.strip()
    a = soup.find("a", href=re.compile(r"/napr/\d+"))
    if a:
        m = re.search(r"\(([\d.]+)\)", a.text)
        if m: data["Код направления"] = m.group(1)

    pi = soup.find("div", class_="podrInfo")
    if pi:
        for d in pi.find_all("div"):
            b = d.find("b")
            if not b: continue
            k = b.text.strip().rstrip(":")
            v = d.text.replace(b.text, "").strip()
            if k in data:
                data[k] = v

    m = re.search(r"(\d+)\s*бюджетн", txt, re.IGNORECASE)
    if m: data["Количество бюджетных мест"] = m.group(1)
    m = re.search(r"(\d+)\s*платн", txt, re.IGNORECASE)
    if m: data["Количество платных мест"] = m.group(1)
    m = re.search(r"(\d[\d\s]+)\s*руб", txt)
    if m: data["Стоимость обучения"] = m.group(1).replace(" ", "")

    sc = parse_scores_from_block(soup)
    for y, val in sc["budget"].items():
        data[f"Проходной балл {y}"] = f"Бюджет: {val}"
    for y, val in sc["paid"].items():
        prev = data[f"Проходной балл {y}"]
        data[f"Проходной балл {y}"] = prev + ("; " if prev else "") + f"Платка: {val}"

    hdr = soup.find("div", string=re.compile("Статистика квот"))
    if hdr:
        divs = hdr.find_all_next("div", class_="col-md-3", limit=4)
        for d in divs:
            cat = d.find("b").text.strip()
            txt = d.get_text(" ", strip=True)
            if "БВИ" in cat:
                m = re.search(r"(\d+)\s*\((\d+)%", txt)
                if m: data["BVI_count"], data["BVI_pct"] = m.groups()
            elif "Целевое" in cat:
                m = re.search(r"(\d+)\s*\((\d+)\)", txt)
                if m: data["Target_count"], data["Target_score"] = m.groups()
            elif "СВО" in cat or "Отдельная" in cat:
                m = re.search(r"(\d+)\s*\((\d+)\)", txt)
                if m: data["SVO_count"], data["SVO_score"] = m.groups()
            elif "Особая" in cat:
                m = re.search(r"(\d+)\s*\((\d+)\)", txt)
                if m: data["Special_count"], data["Special_score"] = m.groups()
    return data

def format_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    ws.freeze_panes = "A2"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align
            cell.border = border
            if cell.row == 1:
                cell.font = header_font

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(path)

def plot_scores(row, fname):
    years = []
    values = []
    for y in range(2020, 2025):
        val = row.get(f"Проходной балл {y}", "")
        m = re.search(r"Бюджет[:\s]*(\d+)", str(val))
        if m:
            years.append(y)
            values.append(int(m.group(1)))
    if len(years) < 2:
        return
    plt.figure(figsize=(5, 3))
    plt.plot(years, values, marker='o', color='blue')
    plt.title(row.get("Название программы", "")[:30])
    plt.xlabel("Год")
    plt.ylabel("Проходной балл (бюджет)")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(fname)
    plt.close()

# --- Главный блок ---
if __name__ == "__main__":
    urls = [
#
# # ТОМСК - ТПУ
#         "https://vuzopedia.ru/vuz/3397/programs/bakispec/2797",  # Прикладная математика в инженерии
#         "https://vuzopedia.ru/vuz/3397/programs/bakispec/5269",  # Информационные технологии и интеллектуальный анализ данных
#         "https://vuzopedia.ru/vuz/3397/programs/bakispec/6728",  # Интеллектуальные системы и технологии
#
#         # ТГУ
#         "https://vuzopedia.ru/vuz/3463/programs/bakispec/3",     # Математическое моделирование и информационные системы
#         "https://vuzopedia.ru/vuz/3463/programs/bakispec/1219",  # Математика в цифровой экономике
#         "https://vuzopedia.ru/vuz/3463/programs/bakispec/2846",  # Искусственный интеллект и большие данные
#
#         # КАЗАНЬ - КФУ
#         "https://vuzopedia.ru/vuz/1751/programs/bakispec/512",   # Математика в цифровой экономике
#         "https://vuzopedia.ru/vuz/1751/programs/bakispec/514",   # Наука о данных
#
#         # КНИТУ
#         "https://vuzopedia.ru/vuz/1773/programs/bakispec/116",   # Информационные системы и технологии
#         "https://vuzopedia.ru/vuz/1773/programs/bakispec/1388",  # Информационные системы и базы данных
#         "https://vuzopedia.ru/vuz/1773/programs/bakispec/3",     # Прикладная математика и информатика
#         "https://vuzopedia.ru/vuz/1773/programs/bakispec/4602",  # Искусственный интеллект и большие данные
#         "https://vuzopedia.ru/vuz/1773/programs/bakispec/2151",  # Бизнес-статистика и анализ данных
#
#         # КГЭУ
#         "https://vuzopedia.ru/vuz/1782/programs/bakispec/6706",  # Технологии разработки цифровых систем и моделей
#         "https://vuzopedia.ru/vuz/1782/programs/bakispec/6295",  # Прикладной искусственный интеллект
#
#         # КГАСУ
#         "https://vuzopedia.ru/vuz/1817/programs/bakispec/2516",  # Информационные системы и технологии в строительстве
#
#         # ЕКАТЕРИНБУРГ - УРФУ
#         "https://vuzopedia.ru/vuz/1848/programs/bakispec/116",   # Информационные системы и технологии
#
#         # МОСКВА - РОСБИОТЕХ
#         "https://vuzopedia.ru/vuz/5154/programs/bakispec/5868",  # Информационные технологии и бизнес-аналитика
#         "https://vuzopedia.ru/vuz/5154/programs/bakispec/3372",  # Искусственный интеллект в управлении технологическими комплексами
#         "https://vuzopedia.ru/vuz/5154/programs/bakispec/828",   # Программно-информационные системы и аналитика
#
#         # РГСУ
#         "https://vuzopedia.ru/vuz/509/programs/bakispec/4619",   # Искусственный интеллект и цифровая гигиена
#         "https://vuzopedia.ru/vuz/509/programs/bakispec/4247",   # Интеллектуальные информационные системы и технологии
#         "https://vuzopedia.ru/vuz/509/programs/bakispec/4623",   # Разработка и внедрение программного обеспечения
#
#         # РГГУ
#         "https://vuzopedia.ru/vuz/96/programs/bakispec/6295",    # Прикладной искусственный интеллект
#         "https://vuzopedia.ru/vuz/96/programs/bakispec/4815",    # Коммуникативные технологии цифровой трансформации
#         "https://vuzopedia.ru/vuz/96/programs/bakispec/7996",    # Математика эффективных ИТ-решений в РГГУ
#         "https://vuzopedia.ru/vuz/96/programs/bakispec/8000",    # Математические основы искусственного интеллекта
#
#         # РГАУ-МСХА им. Тимирязева
#         "https://vuzopedia.ru/vuz/440/programs/bakispec/2782",   # Большие данные и машинное обучение
#         "https://vuzopedia.ru/vuz/440/programs/bakispec/3540",   # Фуллстек разработка
#         "https://vuzopedia.ru/vuz/440/programs/bakispec/3865",   # Системная аналитика и разработка ПО
#         "https://vuzopedia.ru/vuz/440/programs/bakispec/5695",   # Компьютерные науки и технологии искусственного интеллекта
#         "https://vuzopedia.ru/vuz/440/programs/bakispec/2846",   # Системы искусственного интеллекта
#         "https://vuzopedia.ru/vuz/440/programs/bakispec/6096",   # Программные решения для бизнеса
          "https://vuzopedia.ru/vuz/3397/programs/bakispec/9392"

    ]

    print(Fore.CYAN + "🔎 Запускаем парсинг...")
    rows = []
    for url in tqdm(urls, desc="Парсинг"):
        try:
            rows.append(parse_program(url))
        except Exception as e:
            print(Fore.RED + f"Ошибка {url}: {e}")

    df = pd.DataFrame(rows)

    # Графики
    os.makedirs("charts", exist_ok=True)
    for i, r in enumerate(df.to_dict(orient="records"), 1):
        plot_scores(r, f"charts/chart_{i}.png")

    report = "report.xlsx"
    with pd.ExcelWriter(report, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)

    format_excel(report)
    print(Fore.GREEN + "✅ Готово! Отчет: report.xlsx + графики в charts/")
