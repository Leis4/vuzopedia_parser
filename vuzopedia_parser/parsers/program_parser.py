import re
from bs4 import BeautifulSoup
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def parse_scores_from_block(soup):
    scores = {"budget": {}, "paid": {}}
    block = soup.find("div", class_="zpForMabile")
    if not block:
        return scores
    current = None
    for p in block.find_all("p"):
        txt = p.get_text(strip=True).lower()
        if "бюджет" in txt:
            current = "budget"
        elif "плат" in txt:
            current = "paid"
        else:
            m = re.match(r"(\d{4})[:\s]+(\d+)", txt)
            if m and current:
                y, v = m.groups()
                scores[current][y] = int(v)
    return scores


def parse_program(html, url):
    soup = BeautifulSoup(html, 'html.parser')
    data = {}
    data["Ссылка"] = url
    h1 = soup.find("h1")
    if h1:
        data["Название программы"] = h1.text.strip()
    a = soup.find("a", href=re.compile(r"/napr/\d+"))
    if a:
        m = re.search(r"\(([\d.]+)\)", a.text)
        if m:
            data["Код направления"] = m.group(1)
    pi = soup.find("div", class_="podrInfo")
    if pi:
        for d in pi.find_all("div"):
            b = d.find("b")
            if not b: continue
            k = b.text.strip().rstrip(":")
            v = d.get_text(separator=" ", strip=True).replace(b.text, "").strip()
            data[k] = v
    txt = soup.get_text("\n")
    m = re.search(r"(\d+)\s*бюджетн", txt, re.IGNORECASE)
    if m:
        data["Количество бюджетных мест"] = m.group(1)
    m = re.search(r"(\d+)\s*платн", txt, re.IGNORECASE)
    if m:
        data["Количество платных мест"] = m.group(1)
    m = re.search(r"(\d[\d\s]+)\s*руб", txt)
    if m:
        data["Стоимость обучения"] = m.group(1).replace(" ", "")
    sc = parse_scores_from_block(soup)
    for y, val in sc["budget"].items():
        data[f"Проходной балл {y}"] = f"Бюджет: {val}"
    for y, val in sc["paid"].items():
        prev = data.get(f"Проходной балл {y}", "")
        data[f"Проходной балл {y}"] = prev + ("; " if prev else "") + f"Платка: {val}"
    hdr = soup.find("div", string=re.compile("Статистика квот", re.IGNORECASE))
    if hdr:
        divs = hdr.find_all_next("div", class_="col-md-3", limit=4)
        for d in divs:
            b = d.find("b")
            if not b: continue
            cat = b.text.strip()
            txt_div = d.get_text(" ", strip=True)
            if "БВИ" in cat:
                m = re.search(r"(\d+)\s*\((\d+)%", txt_div)
                if m: data["BVI_count"], data["BVI_pct"] = m.groups()
            elif "Целевое" in cat:
                m = re.search(r"(\d+)\s*\((\d+)\)", txt_div)
                if m: data["Target_count"], data["Target_score"] = m.groups()
            elif "СВО" in cat or "Отдельная" in cat:
                m = re.search(r"(\d+)\s*\((\d+)\)", txt_div)
                if m: data["SVO_count"], data["SVO_score"] = m.groups()
            elif "Особая" in cat:
                m = re.search(r"(\d+)\s*\((\d+)\)", txt_div)
                if m: data["Special_count"], data["Special_score"] = m.groups()
    return data


def format_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align
            cell.border = border
            if cell.row == 1:
                cell.font = header_font
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_len + 2
    wb.save(path)
